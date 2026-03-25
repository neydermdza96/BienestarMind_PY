"""
Generador de Reportes — BienestarMind
PDF con ReportLab · Excel con openpyxl
Paleta SENA: Verde #3d9b35 / Naranja #f47920
"""
import io
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, Image,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing, Rect
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Paleta SENA ────────────────────────────────────────────────────────────
VERDE_SENA = colors.HexColor('#3d9b35')
NARANJA_SENA = colors.HexColor('#f47920')
GRIS_CLARO = colors.HexColor('#f5f5f5')
GRIS_OSCURO = colors.HexColor('#333333')
BLANCO = colors.white

VERDE_XL = 'FF3d9b35'
NARANJA_XL = 'FFf47920'
BLANCO_XL = 'FFFFFFFF'
GRIS_XL = 'FFf5f5f5'
TEXTO_XL = 'FF333333'


# ── PDF ────────────────────────────────────────────────────────────────────

def _encabezado_pdf(story, titulo, subtitulo=''):
    styles = getSampleStyleSheet()

    titulo_style = ParagraphStyle(
        'TituloSENA', parent=styles['Title'],
        textColor=VERDE_SENA, fontSize=18, spaceAfter=4,
        fontName='Helvetica-Bold',
    )
    sub_style = ParagraphStyle(
        'SubSENA', parent=styles['Normal'],
        textColor=NARANJA_SENA, fontSize=11, spaceAfter=2,
        fontName='Helvetica-Bold',
    )
    meta_style = ParagraphStyle(
        'MetaSENA', parent=styles['Normal'],
        textColor=GRIS_OSCURO, fontSize=9,
        fontName='Helvetica',
    )

    story.append(Paragraph('🧠 BienestarMind', titulo_style))
    story.append(Paragraph('Servicio Nacional de Aprendizaje — SENA', sub_style))
    story.append(Paragraph(f'Reporte: {titulo}', ParagraphStyle('t2', parent=styles['Normal'], fontSize=13, textColor=GRIS_OSCURO, fontName='Helvetica-Bold')))
    if subtitulo:
        story.append(Paragraph(subtitulo, meta_style))
    story.append(Paragraph(f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', meta_style))
    story.append(HRFlowable(width='100%', thickness=2, color=VERDE_SENA, spaceAfter=10))


def _tabla_pdf(headers, rows, col_widths=None):
    data = [headers] + rows
    tabla = Table(data, colWidths=col_widths, repeatRows=1)
    tabla.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), VERDE_SENA),
        ('TEXTCOLOR', (0, 0), (-1, 0), BLANCO),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        # Filas alternas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [BLANCO, GRIS_CLARO]),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('LINEBELOW', (0, 0), (-1, 0), 1.5, NARANJA_SENA),
    ]))
    return tabla


def _pie_pdf(story, total):
    styles = getSampleStyleSheet()
    pie_style = ParagraphStyle('pie', parent=styles['Normal'], fontSize=8,
                                textColor=colors.HexColor('#888888'), alignment=TA_CENTER)
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width='100%', thickness=1, color=NARANJA_SENA))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f'Total de registros: {total}  |  BienestarMind SENA  |  Documento generado automáticamente', pie_style))


def generar_pdf_asesorias(asesorias):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    _encabezado_pdf(story, 'Asesorías Registradas')

    headers = ['#', 'Fecha', 'Asesor', 'Aprendiz', 'Ficha', 'Motivo', 'Estado']
    rows = []
    for a in asesorias:
        rows.append([
            str(a.pk),
            a.fecha.strftime('%d/%m/%Y') if a.fecha else '-',
            a.usuario_asesor.nombre_completo if a.usuario_asesor else '-',
            a.usuario_recibe.nombre_completo if a.usuario_recibe else '-',
            str(a.ficha.id_ficha) if a.ficha else '-',
            (a.motivo_asesoria[:60] + '...') if len(a.motivo_asesoria) > 60 else a.motivo_asesoria,
            a.get_estado_display(),
        ])

    story.append(_tabla_pdf(headers, rows, [1*cm, 2.2*cm, 3.5*cm, 3.5*cm, 2*cm, 8*cm, 2.5*cm]))
    _pie_pdf(story, len(rows))
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_pdf_reservas_espacios(reservas):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    _encabezado_pdf(story, 'Reservas de Espacios')

    headers = ['#', 'Fecha', 'Espacio', 'Sede', 'Solicitante', 'Duración', 'Motivo', 'Estado']
    rows = []
    for r in reservas:
        rows.append([
            str(r.pk),
            r.fecha_reserva.strftime('%d/%m/%Y'),
            r.espacio.nombre_del_espacio,
            r.espacio.sede.nombre_sede,
            r.usuario.nombre_completo,
            f'{r.duracion} min',
            (r.motivo_reserva[:50] + '...') if len(r.motivo_reserva) > 50 else r.motivo_reserva,
            r.get_estado_display(),
        ])

    story.append(_tabla_pdf(headers, rows, [0.8*cm, 2.2*cm, 3.5*cm, 3*cm, 3.5*cm, 1.8*cm, 6*cm, 2.2*cm]))
    _pie_pdf(story, len(rows))
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_pdf_reservas_elementos(reservas):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    _encabezado_pdf(story, 'Reservas de Elementos')

    headers = ['#', 'Fecha', 'Elemento', 'Categoría', 'Solicitante', 'Estado']
    rows = []
    for r in reservas:
        rows.append([
            str(r.pk),
            r.fecha_reserva.strftime('%d/%m/%Y'),
            r.elemento.nombre_elemento,
            r.elemento.categoria.descripcion,
            r.usuario.nombre_completo,
            r.get_estado_display(),
        ])

    story.append(_tabla_pdf(headers, rows))
    _pie_pdf(story, len(rows))
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_pdf_usuarios(usuarios):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    _encabezado_pdf(story, 'Directorio de Usuarios')

    headers = ['Nombres', 'Apellidos', 'Documento', 'Correo', 'Teléfono', 'Género', 'Edad', 'Roles']
    rows = []
    for u in usuarios:
        roles = ', '.join([r.descripcion for r in u.roles.all()])
        rows.append([
            u.nombres, u.apellidos, u.documento,
            u.correo, u.telefono, u.get_genero_display(),
            f'{u.edad} años', roles,
        ])

    story.append(_tabla_pdf(headers, rows))
    _pie_pdf(story, len(rows))
    doc.build(story)
    buffer.seek(0)
    return buffer


# ── EXCEL ──────────────────────────────────────────────────────────────────

def _setup_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    return wb, ws


def _header_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=BLANCO_XL, size=10)
    cell.fill = PatternFill(start_color=VERDE_XL, end_color=VERDE_XL, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        bottom=Side(style='medium', color=NARANJA_XL),
        right=Side(style='thin', color='FFcccccc'),
    )
    return cell


def _data_cell(ws, row, col, value, alt=False):
    cell = ws.cell(row=row, column=col, value=value)
    bg = GRIS_XL if alt else BLANCO_XL
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    cell.border = Border(right=Side(style='thin', color='FFdddddd'), bottom=Side(style='thin', color='FFeeeeee'))
    return cell


def _encabezado_excel(ws, titulo):
    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value = f'BienestarMind SENA — {titulo}'
    t.font = Font(bold=True, size=14, color=VERDE_XL)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:H2')
    s = ws['A2']
    s.value = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Servicio Nacional de Aprendizaje'
    s.font = Font(size=9, color=NARANJA_XL, italic=True)
    s.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16


def generar_excel_asesorias(asesorias):
    wb, ws = _setup_excel()
    ws.title = 'Asesorías'
    _encabezado_excel(ws, 'Asesorías Registradas')

    headers = ['#', 'Fecha', 'Asesor', 'Aprendiz', 'Ficha', 'Motivo', 'Estado', 'Notif. Enviada']
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 3, col, h)
    ws.row_dimensions[3].height = 22

    for i, a in enumerate(asesorias, 4):
        alt = i % 2 == 0
        data = [
            a.pk, a.fecha.strftime('%d/%m/%Y') if a.fecha else '-',
            a.usuario_asesor.nombre_completo if a.usuario_asesor else '-',
            a.usuario_recibe.nombre_completo if a.usuario_recibe else '-',
            a.ficha.id_ficha if a.ficha else '-',
            a.motivo_asesoria,
            a.get_estado_display(),
            'Sí' if a.notificacion_enviada else 'No',
        ]
        for col, val in enumerate(data, 1):
            _data_cell(ws, i, col, val, alt)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:H{3 + len(asesorias)}'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_excel_reservas_espacios(reservas):
    wb, ws = _setup_excel()
    ws.title = 'Reservas Espacios'
    _encabezado_excel(ws, 'Reservas de Espacios')

    headers = ['#', 'Fecha', 'Espacio', 'Sede', 'Solicitante', 'Duración (min)', 'Motivo', 'Estado']
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 3, col, h)

    for i, r in enumerate(reservas, 4):
        alt = i % 2 == 0
        data = [
            r.pk, r.fecha_reserva.strftime('%d/%m/%Y'),
            r.espacio.nombre_del_espacio, r.espacio.sede.nombre_sede,
            r.usuario.nombre_completo, r.duracion,
            r.motivo_reserva, r.get_estado_display(),
        ]
        for col, val in enumerate(data, 1):
            _data_cell(ws, i, col, val, alt)

    for col, w in enumerate([5, 14, 22, 18, 22, 14, 35, 14], 1):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = w
    ws.freeze_panes = 'A4'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_excel_usuarios(usuarios):
    wb, ws = _setup_excel()
    ws.title = 'Usuarios'
    _encabezado_excel(ws, 'Directorio de Usuarios')

    headers = ['Nombres', 'Apellidos', 'Documento', 'Correo', 'Teléfono', 'Género', 'Edad', 'Roles']
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 3, col, h)

    for i, u in enumerate(usuarios, 4):
        alt = i % 2 == 0
        roles = ', '.join([r.descripcion for r in u.roles.all()])
        data = [u.nombres, u.apellidos, u.documento, u.correo, u.telefono, u.get_genero_display(), u.edad, roles]
        for col, val in enumerate(data, 1):
            _data_cell(ws, i, col, val, alt)

    for col, w in enumerate([20, 20, 16, 28, 16, 14, 8, 24], 1):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = w
    ws.freeze_panes = 'A4'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
